"""
Import Excel de produits, avec prix de gros et de détail.

Le modèle est passé en version 1.3 : prix d'achat au conditionnement, mode de
vente saisi plutôt que deviné, en-têtes de prix qui disent gros ou détail. Les
fichiers de la version 1.2, téléchargés avant ce changement, doivent continuer
de passer : un marchand peut avoir préparé son fichier depuis des semaines.
"""
import io
from decimal import Decimal

from openpyxl import Workbook, load_workbook
from rest_framework.test import APITestCase

from apps.products.models import Product, Unit
from apps.products.services import ProductExcelService
from apps.sales.tests._helpers import make_org_with_users


def _workbook_bytes(headers, rows):
    """Classeur minimal au format attendu : en-têtes en ligne 2, données en 3."""
    wb = Workbook()
    ws = wb.active
    ws.title = "📦 Produits"
    ws.cell(row=1, column=1, value="VF|test|0000000000000000")
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=2, column=col_idx, value=header)
    for row_idx, row in enumerate(rows, start=3):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


class _ImportSetup(APITestCase):

    def setUp(self):
        ctx = make_org_with_users()
        self.__dict__.update(ctx)
        self.headers = [col["header"] for col in ProductExcelService.COLUMNS]
        self.keys = [col["key"] for col in ProductExcelService.COLUMNS]

    def _row(self, **values):
        """Ligne complète dans l'ordre des colonnes courantes."""
        return [values.get(key) for key in self.keys]

    def _import(self, rows, headers=None):
        return ProductExcelService.import_products(
            _workbook_bytes(headers or self.headers, rows),
            self.org,
            self.owner,
        )

    def _errors(self, result):
        return [message for error in result["errors"] for message in error["errors"]]


class TemplateGenerationTests(_ImportSetup):

    def test_template_contient_les_quatre_prix_et_le_mode_de_vente(self):
        buffer = ProductExcelService.generate_template(self.org)
        ws = load_workbook(buffer)["📦 Produits"]

        headers = [ws.cell(row=2, column=i).value for i in range(1, len(self.headers) + 1)]
        for expected in (
            "Mode de vente",
            "Prix d'achat (détail)",
            "Prix de vente (détail) *",
            "Prix d'achat (conditionnement)",
            "Prix de vente (conditionnement)",
        ):
            self.assertIn(expected, headers)

    def test_template_genere_est_reimportable(self):
        """
        Aller-retour : le fichier que le marchand télécharge doit passer la
        validation d'import. C'est ce qui verrouille toute coquille d'en-tête.
        """
        buffer = ProductExcelService.generate_template(self.org)
        is_valid, message, _wb = ProductExcelService.validate_import_file(
            buffer.getvalue(), self.org
        )
        self.assertTrue(is_valid, message)


class HeaderCompatibilityTests(_ImportSetup):

    def test_ancien_modele_v12_est_encore_accepte(self):
        legacy = ProductExcelService.LEGACY_LAYOUTS["1.2"]
        legacy_headers = [header for _key, header in legacy]
        row = [None] * len(legacy)
        for index, (key, _header) in enumerate(legacy):
            row[index] = {
                'name': 'Savon', 'sku': 'SAV-01', 'selling_price': 800,
                'cost_price': 400,
            }.get(key)

        result = self._import([row], headers=legacy_headers)

        self.assertEqual(result["created"], 1, self._errors(result))
        self.assertEqual(Product.objects.get(sku='SAV-01').selling_price, Decimal('800'))

    def test_en_tetes_inconnus_sont_refuses(self):
        result = self._import(
            [['Savon', 'SAV-01']], headers=['Nom bricolé', 'Autre chose']
        )

        self.assertFalse(result["success"])
        self.assertIn("modèle", result["error"])
        self.assertEqual(result["created"], 0)


class PackagingImportTests(_ImportSetup):

    def _packaged_row(self, **overrides):
        values = {
            'name': 'Eau 50cl', 'sku': 'EAU-50',
            'unit': 'Bouteille', 'packaging_unit': 'Carton',
            'selling_mode': 'Gros et détail', 'units_per_package': 12,
            'selling_price': 550, 'wholesale_price': 6000,
            'package_cost_price': 5400,
        }
        values.update(overrides)
        return self._row(**values)

    def test_unite_de_gros_inconnue_est_creee(self):
        """
        Le carton n'existe pas encore : il doit être créé, sinon le produit
        retombait silencieusement en vente au détail seule.
        """
        result = self._import([self._packaged_row()])

        self.assertEqual(result["created"], 1, self._errors(result))
        product = Product.objects.get(sku='EAU-50')
        self.assertEqual(product.selling_mode, Product.SellingMode.WHOLESALE_AND_RETAIL)
        self.assertIsNotNone(product.packaging_unit)
        self.assertEqual(product.packaging_unit.name, 'Carton')
        self.assertTrue(Unit.objects.filter(organization=self.org, name='Carton').exists())

    def test_prix_dachat_unitaire_deduit_du_conditionnement(self):
        self._import([self._packaged_row()])

        product = Product.objects.get(sku='EAU-50')
        self.assertEqual(product.package_cost_price, Decimal('5400.00'))
        self.assertEqual(product.cost_price, Decimal('450.00'))

    def test_prix_dachat_unitaire_saisi_prime(self):
        self._import([self._packaged_row(cost_price=470)])

        self.assertEqual(Product.objects.get(sku='EAU-50').cost_price, Decimal('470'))

    def test_mode_gros_seul_est_atteignable(self):
        """Inatteignable avant la 1.3, où le mode était déduit."""
        result = self._import([self._packaged_row(
            selling_mode='Gros', selling_price=None
        )])

        self.assertEqual(result["created"], 1, self._errors(result))
        self.assertEqual(
            Product.objects.get(sku='EAU-50').selling_mode,
            Product.SellingMode.WHOLESALE_ONLY,
        )

    def test_libelles_de_mode_tolerants(self):
        rows = [
            self._packaged_row(sku='EAU-1', selling_mode='GROS ET DETAIL'),
            self._packaged_row(sku='EAU-2', selling_mode='  les deux  '),
            self._packaged_row(sku='EAU-3', selling_mode='wholesale_and_retail'),
        ]
        result = self._import(rows)

        self.assertEqual(result["created"], 3, self._errors(result))
        for sku in ('EAU-1', 'EAU-2', 'EAU-3'):
            self.assertEqual(
                Product.objects.get(sku=sku).selling_mode,
                Product.SellingMode.WHOLESALE_AND_RETAIL,
            )

    def test_mode_inconnu_est_rejete(self):
        result = self._import([self._packaged_row(selling_mode='en vrac')])

        self.assertEqual(result["created"], 0)
        self.assertEqual(result["skipped"], 1)
        self.assertIn("Mode de vente non reconnu", " ".join(self._errors(result)))

    def test_ligne_gros_incomplete_est_rejetee_avec_ses_manques(self):
        """
        Plus de dégradation silencieuse : sans ce rejet, la contrainte de base
        ferait échouer la création de tout le lot.
        """
        result = self._import([
            self._packaged_row(
                sku='EAU-KO', packaging_unit=None, units_per_package=1,
                wholesale_price=None,
            ),
            self._packaged_row(sku='EAU-OK'),
        ])

        self.assertEqual(result["created"], 1, self._errors(result))
        self.assertEqual(result["skipped"], 1)
        messages = " ".join(self._errors(result))
        self.assertIn("Unité de gros manquante", messages)
        self.assertIn("au moins 2", messages)
        self.assertIn("Prix de vente (conditionnement) manquant", messages)
        self.assertTrue(Product.objects.filter(sku='EAU-OK').exists())

    def test_colonne_mode_vide_deduit_comme_avant(self):
        result = self._import([self._packaged_row(selling_mode=None)])

        self.assertEqual(result["created"], 1, self._errors(result))
        self.assertEqual(
            Product.objects.get(sku='EAU-50').selling_mode,
            Product.SellingMode.WHOLESALE_AND_RETAIL,
        )


class PriceValidationTests(_ImportSetup):

    def test_prix_illisible_est_rejete_au_lieu_de_valoir_zero(self):
        result = self._import([self._row(
            name='Savon', sku='SAV-01', selling_price=800, cost_price='6 000 FC'
        )])

        self.assertEqual(result["created"], 0)
        self.assertIn("illisible", " ".join(self._errors(result)))

    def test_prix_de_vente_inferieur_au_prix_dachat_est_rejete(self):
        result = self._import([self._row(
            name='Savon', sku='SAV-01', selling_price=300, cost_price=400
        )])

        self.assertEqual(result["created"], 0)
        self.assertIn(
            "ne peut pas être inférieur", " ".join(self._errors(result))
        )

    def test_prix_de_vente_detail_requis_hors_mode_gros(self):
        result = self._import([self._row(name='Savon', sku='SAV-01')])

        self.assertEqual(result["created"], 0)
        self.assertIn("Prix de vente (détail) requis", self._errors(result))

    def test_import_au_detail_reste_inchange(self):
        """Non-régression du cas le plus courant."""
        result = self._import([self._row(
            name='Savon', sku='SAV-01', unit='Pièce', category='Hygiène',
            brand='Palmolive', cost_price=400, selling_price=800, tax_rate=16,
            is_taxable='Oui', min_stock_level=5,
        )])

        self.assertEqual(result["created"], 1, self._errors(result))
        product = Product.objects.get(sku='SAV-01')
        self.assertEqual(product.selling_mode, Product.SellingMode.RETAIL_ONLY)
        self.assertIsNone(product.units_per_package)
        self.assertEqual(product.cost_price, Decimal('400'))
        self.assertEqual(product.tax_rate, Decimal('16'))


class ImportGuardTests(_ImportSetup):
    """Comportements existants, verrouillés maintenant qu'on touche au fichier."""

    def test_sku_duplique_est_ignore(self):
        Product.objects.create(
            organization=self.org, name='Savon', slug='savon', sku='SAV-01',
            cost_price=Decimal('400'), selling_price=Decimal('800'),
        )
        result = self._import([self._row(
            name='Savon bis', sku='SAV-01', selling_price=900
        )])

        self.assertEqual(result["created"], 0)
        self.assertEqual(result["skipped"], 1)
        self.assertIn("existe déjà", " ".join(self._errors(result)))

    def test_sku_manquant_est_ignore(self):
        result = self._import([self._row(name='Savon', selling_price=800)])

        self.assertEqual(result["created"], 0)
        self.assertIn("Code SKU requis", self._errors(result))
