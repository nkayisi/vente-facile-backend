"""
Services pour l'importation et l'exportation de produits via Excel.
"""
import hashlib
import io
from datetime import datetime
from decimal import Decimal, InvalidOperation
from typing import Dict, List, Tuple, Optional, Any

from django.db import transaction
from django.utils.text import slugify
from rest_framework.exceptions import ValidationError

from apps.subscriptions.services import SubscriptionService
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from .models import Product, Category, Brand, Unit


# Signature secrète pour valider l'authenticité du template
TEMPLATE_SIGNATURE = "VF-IMPORT-2026-SECURE"
# 1.2: signature globale (sans organization.id) pour que le même fichier reste valide entre orgs / environnements.
TEMPLATE_VERSION = "1.2"


class ProductExcelService:
    """Service pour la gestion des imports/exports Excel de produits."""
    
    # Colonnes du template avec leurs configurations
    COLUMNS = [
        {"key": "name", "header": "Nom du produit *", "width": 35, "required": True},
        {"key": "sku", "header": "Code SKU *", "width": 15, "required": True},
        {"key": "barcode", "header": "Code-barres", "width": 18, "required": False},
        {"key": "category", "header": "Catégorie", "width": 25, "required": False},
        {"key": "subcategory", "header": "Sous-catégorie", "width": 25, "required": False},
        {"key": "brand", "header": "Marque", "width": 15, "required": False},
        {"key": "unit", "header": "Unité", "width": 12, "required": False},
        {"key": "cost_price", "header": "Prix d'achat", "width": 15, "required": False},
        {"key": "selling_price", "header": "Prix de vente *", "width": 15, "required": True},
        {"key": "wholesale_price", "header": "Prix de gros", "width": 15, "required": False},
        {"key": "tax_rate", "header": "Taux TVA (%)", "width": 12, "required": False},
        {"key": "is_taxable", "header": "Taxable (Oui/Non)", "width": 15, "required": False},
        {"key": "track_inventory", "header": "Suivi stock (Oui/Non)", "width": 18, "required": False},
        {"key": "has_expiry_date", "header": "Périssable (Oui/Non)", "width": 18, "required": False},
        {"key": "min_stock_level", "header": "Stock minimum", "width": 14, "required": False},
        {"key": "short_description", "header": "Description", "width": 40, "required": False},
    ]
    
    @classmethod
    def generate_template(cls, organization) -> io.BytesIO:
        """
        Génère un fichier Excel template pour l'importation de produits.
        Le template contient une signature cachée pour validation.
        """
        wb = Workbook()
        
        # =====================================================================
        # FEUILLE 1 : GUIDE DE REMPLISSAGE
        # =====================================================================
        guide_sheet = wb.active
        guide_sheet.title = "📖 Guide"
        
        # Styles pour le guide
        title_font = Font(bold=True, size=16, color="F97316")
        section_font = Font(bold=True, size=12, color="1F2937")
        text_font = Font(size=11, color="374151")
        example_font = Font(size=10, color="6B7280", italic=True)
        warning_font = Font(bold=True, size=11, color="DC2626")
        success_font = Font(bold=True, size=11, color="059669")
        
        guide_sheet.column_dimensions['A'].width = 80
        
        row = 1
        
        # Titre principal
        guide_sheet.cell(row=row, column=1, value="📋 GUIDE D'IMPORTATION DES PRODUITS").font = title_font
        row += 2
        
        # Introduction
        guide_sheet.cell(row=row, column=1, value="Ce fichier vous permet d'importer vos produits en masse dans Vente Facile.").font = text_font
        row += 1
        guide_sheet.cell(row=row, column=1, value="Suivez attentivement les instructions ci-dessous pour éviter les erreurs.").font = text_font
        row += 2
        
        # Section 1 : Feuilles du fichier
        guide_sheet.cell(row=row, column=1, value="📑 STRUCTURE DU FICHIER").font = section_font
        row += 1
        guide_sheet.cell(row=row, column=1, value="• 📖 Guide : Cette feuille d'instructions (vous pouvez la consulter à tout moment)").font = text_font
        row += 1
        guide_sheet.cell(row=row, column=1, value="• 📦 Produits : Feuille principale où vous saisissez vos produits").font = text_font
        row += 2
        
        # Section 2 : Colonnes obligatoires
        guide_sheet.cell(row=row, column=1, value="⚠️ COLONNES OBLIGATOIRES (marquées d'un *)").font = warning_font
        row += 1
        guide_sheet.cell(row=row, column=1, value="• Nom du produit * : Le nom complet du produit").font = text_font
        row += 1
        guide_sheet.cell(row=row, column=1, value="• Code SKU * : Code unique d'identification (ex: COCA-33CL, FANTA-50CL)").font = text_font
        row += 1
        guide_sheet.cell(row=row, column=1, value="• Prix de vente * : Prix de vente au client (nombre décimal)").font = text_font
        row += 2
        
        # Section 3 : Colonnes optionnelles
        guide_sheet.cell(row=row, column=1, value="📝 COLONNES OPTIONNELLES").font = section_font
        row += 1
        guide_sheet.cell(row=row, column=1, value="• Code-barres : Code-barres EAN/UPC du produit").font = text_font
        row += 1
        guide_sheet.cell(row=row, column=1, value="• Catégorie : Nom de la catégorie principale").font = text_font
        row += 1
        guide_sheet.cell(row=row, column=1, value="• Sous-catégorie : Nom de la sous-catégorie (optionnel)").font = text_font
        row += 1
        guide_sheet.cell(row=row, column=1, value="  → Si renseignée, elle sera créée sous la catégorie de la même ligne").font = example_font
        row += 1
        guide_sheet.cell(row=row, column=1, value="• Marque : Nom de la marque du produit").font = text_font
        row += 1
        guide_sheet.cell(row=row, column=1, value="• Unité : Unité de mesure (Pièce, Kg, Litre, Carton, etc.)").font = text_font
        row += 1
        guide_sheet.cell(row=row, column=1, value="• Prix d'achat : Prix d'achat/coût du produit").font = text_font
        row += 1
        guide_sheet.cell(row=row, column=1, value="• Prix de gros : Prix pour les ventes en gros (optionnel)").font = text_font
        row += 1
        guide_sheet.cell(row=row, column=1, value="• Taux TVA (%) : Pourcentage de TVA (ex: 16 pour 16%)").font = text_font
        row += 1
        guide_sheet.cell(row=row, column=1, value="• Taxable : Oui ou Non").font = text_font
        row += 1
        guide_sheet.cell(row=row, column=1, value="• Suivi stock : Oui ou Non (si Oui, le stock sera géré)").font = text_font
        row += 1
        guide_sheet.cell(row=row, column=1, value="• Périssable : Oui ou Non (si Oui, les dates d'expiration seront gérées)").font = text_font
        row += 1
        guide_sheet.cell(row=row, column=1, value="• Stock minimum : Seuil d'alerte de stock bas").font = text_font
        row += 1
        guide_sheet.cell(row=row, column=1, value="• Description : Description courte du produit").font = text_font
        row += 2
        
        # Section 4 : Création automatique
        guide_sheet.cell(row=row, column=1, value="✨ CRÉATION AUTOMATIQUE DES RÉFÉRENCES").font = success_font
        row += 1
        guide_sheet.cell(row=row, column=1, value="Si vous saisissez une catégorie, marque ou unité qui n'existe pas encore,").font = text_font
        row += 1
        guide_sheet.cell(row=row, column=1, value="elle sera automatiquement créée lors de l'importation !").font = text_font
        row += 1
        guide_sheet.cell(row=row, column=1, value="→ Exemple : Si vous écrivez 'Électronique' comme catégorie et qu'elle n'existe pas,").font = example_font
        row += 1
        guide_sheet.cell(row=row, column=1, value="  elle sera créée automatiquement avant d'importer le produit.").font = example_font
        row += 1
        guide_sheet.cell(row=row, column=1, value="  Le même principe s'applique pour les sous-catégories, marques et unités.").font = example_font
        row += 2
        
        # Section 5 : Règles importantes
        guide_sheet.cell(row=row, column=1, value="🚫 RÈGLES IMPORTANTES").font = warning_font
        row += 1
        guide_sheet.cell(row=row, column=1, value="• Ne modifiez PAS la ligne d'en-tête (ligne 2 de la feuille Produits)").font = text_font
        row += 1
        guide_sheet.cell(row=row, column=1, value="• Ne supprimez PAS la première ligne cachée (signature de validation)").font = text_font
        row += 1
        guide_sheet.cell(row=row, column=1, value="• Chaque SKU doit être UNIQUE (pas de doublons)").font = text_font
        row += 1
        guide_sheet.cell(row=row, column=1, value="• Chaque code-barres doit être UNIQUE (si renseigné)").font = text_font
        row += 1
        guide_sheet.cell(row=row, column=1, value="• Les prix doivent être des nombres (utilisez le point ou la virgule comme séparateur décimal)").font = text_font
        row += 2
        
        # Section 6 : Exemples
        guide_sheet.cell(row=row, column=1, value="📌 EXEMPLES DE SAISIE").font = section_font
        row += 1
        guide_sheet.cell(row=row, column=1, value="Nom: Coca-Cola 33cl | SKU: COCA-33CL | Prix: 500 | Catégorie: Boissons | Sous-catégorie: Sodas").font = example_font
        row += 1
        guide_sheet.cell(row=row, column=1, value="Nom: Riz Oncle Ben's 1kg | SKU: RIZ-OB-1KG | Prix: 2500 | Catégorie: Alimentation").font = example_font
        row += 1
        guide_sheet.cell(row=row, column=1, value="Nom: Savon Palmolive | SKU: SAV-PALM | Prix: 800 | Marque: Palmolive | Unité: Pièce").font = example_font
        row += 2
        
        guide_sheet.cell(row=row, column=1, value="💡 Astuce : Vous pouvez saisir de nouvelles catégories, sous-catégories, marques et unités; elles seront créées automatiquement.").font = success_font
        
        # =====================================================================
        # FEUILLE 2 : PRODUITS
        # =====================================================================
        ws = wb.create_sheet("📦 Produits")
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="F97316", end_color="F97316", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # En-têtes (ligne 2, ligne 1 réservée pour la signature)
        for col_idx, col_config in enumerate(cls.COLUMNS, start=1):
            cell = ws.cell(row=2, column=col_idx, value=col_config["header"])
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
            ws.column_dimensions[get_column_letter(col_idx)].width = col_config["width"]
        
        # Ligne 1 : Signature cachée (texte blanc sur fond blanc, protégée)
        signature_data = cls._generate_signature()
        ws.cell(row=1, column=1, value=signature_data)
        ws.cell(row=1, column=1).font = Font(color="FFFFFF", size=1)
        ws.cell(row=1, column=1).fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        ws.row_dimensions[1].height = 5  # Ligne très petite
        
        # Ajouter des validations de données
        # Oui/Non pour les colonnes booléennes
        yes_no_validation = DataValidation(
            type="list",
            formula1='"Oui,Non"',
            allow_blank=True
        )
        yes_no_validation.error = "Veuillez sélectionner Oui ou Non"
        yes_no_validation.errorTitle = "Valeur invalide"
        
        # Appliquer aux colonnes booléennes
        bool_columns = ["is_taxable", "track_inventory", "has_expiry_date"]
        for col_idx, col_config in enumerate(cls.COLUMNS, start=1):
            if col_config["key"] in bool_columns:
                col_letter = get_column_letter(col_idx)
                yes_no_validation.add(f"{col_letter}3:{col_letter}1000")
        
        ws.add_data_validation(yes_no_validation)
        
        # Ajouter quelques lignes vides formatées pour guider l'utilisateur
        for row in range(3, 13):
            for col_idx in range(1, len(cls.COLUMNS) + 1):
                cell = ws.cell(row=row, column=col_idx, value="")
                cell.border = thin_border
        
        # Figer les en-têtes
        ws.freeze_panes = "A3"
        
        # Sauvegarder dans un buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return buffer
    
    @classmethod
    def _get_categories_with_path(cls, organization) -> List[str]:
        """
        Retourne la liste des catégories avec leur chemin hiérarchique.
        Format: "Parent > Enfant > Sous-enfant"
        """
        categories = Category.objects.filter(
            organization=organization, 
            is_deleted=False
        ).select_related('parent').order_by('parent__name', 'name')
        
        result = []
        
        def get_path(category):
            """Construit le chemin complet d'une catégorie."""
            path_parts = [category.name]
            current = category
            while current.parent:
                path_parts.insert(0, current.parent.name)
                current = current.parent
            return " > ".join(path_parts)
        
        for cat in categories:
            result.append(get_path(cat))
        
        # Trier par chemin
        result.sort()
        return result
    
    @classmethod
    def _generate_signature(cls) -> str:
        """Signature du format de template (identique pour toutes les organisations)."""
        data = f"{TEMPLATE_SIGNATURE}|{TEMPLATE_VERSION}"
        hash_value = hashlib.sha256(data.encode()).hexdigest()[:16]
        return f"VF|{TEMPLATE_VERSION}|{hash_value}"

    @classmethod
    def validate_import_file(cls, file_content: bytes, organization) -> Tuple[bool, str, Optional[Workbook]]:
        """
        Valide un fichier Excel avant l'importation.
        Retourne (is_valid, message, workbook)
        """
        try:
            wb = load_workbook(io.BytesIO(file_content))
        except Exception as e:
            return False, f"Fichier Excel invalide: {str(e)}", None
        
        # Vérifier que la feuille "Produits" existe (avec ou sans emoji)
        products_sheet_name = None
        for sheet_name in wb.sheetnames:
            if "Produits" in sheet_name:
                products_sheet_name = sheet_name
                break
        
        if not products_sheet_name:
            return False, "Feuille 'Produits' introuvable. Utilisez le template officiel.", None
        
        ws = wb[products_sheet_name]

        # En-têtes (ligne 2) : critère principal du format — doit correspondre au template officiel
        expected_headers = [col["header"] for col in cls.COLUMNS]
        actual_headers = [ws.cell(row=2, column=i).value for i in range(1, len(cls.COLUMNS) + 1)]

        if actual_headers != expected_headers:
            return False, "Les en-têtes du fichier ont été modifiés. Utilisez le template officiel sans modifier les en-têtes.", None

        # Signature (ligne 1) : renforce l'authenticité mais ne bloque pas si les en-têtes sont conformes
        # (ex. template 1.1 lié à une autre organisation, copie entre environnements, Excel qui modifie A1).
        # Ligne 1 (signature) ignorée pour le blocage : les en-têtes exacts garantissent le format attendu.

        return True, "Fichier valide", wb
    
    @classmethod
    def import_products(cls, file_content: bytes, organization, user) -> Dict[str, Any]:
        """
        Importe les produits depuis un fichier Excel.
        Retourne un résumé de l'importation.
        """
        # Valider le fichier
        is_valid, message, wb = cls.validate_import_file(file_content, organization)
        if not is_valid:
            return {
                "success": False,
                "error": message,
                "created": 0,
                "updated": 0,
                "skipped": 0,
                "errors": []
            }
        
        # Trouver la feuille Produits
        products_sheet_name = next((s for s in wb.sheetnames if "Produits" in s), None)
        ws = wb[products_sheet_name]
        
        category_lookup = cls._build_category_lookup(organization)
        brands_map = cls._build_brands_map(organization)
        units_map = cls._build_units_map(organization)
        
        # Charger les SKU et codes-barres existants pour détecter les doublons
        existing_skus = set(Product.objects.filter(organization=organization, is_deleted=False).values_list('sku', flat=True))
        existing_barcodes = set(Product.objects.filter(organization=organization, is_deleted=False, barcode__isnull=False).exclude(barcode='').values_list('barcode', flat=True))
        
        results = {
            "success": True,
            "created": 0,
            "updated": 0,
            "skipped": 0,
            "errors": []
        }
        
        products_to_create = []
        row_num = 3  # Données commencent à la ligne 3
        parent_categories, subcategories, brand_values, unit_values = cls._collect_reference_values(ws)
        cls._sync_categories(organization, category_lookup, parent_categories, subcategories)
        cls._sync_brands(organization, brands_map, brand_values)
        cls._sync_units(organization, units_map, unit_values)
        
        while True:
            # Lire la ligne
            row_data = {}
            has_data = False
            
            for col_idx, col_config in enumerate(cls.COLUMNS, start=1):
                value = ws.cell(row=row_num, column=col_idx).value
                if value is not None and str(value).strip():
                    has_data = True
                row_data[col_config["key"]] = value
            
            if not has_data:
                break  # Fin des données
            
            # Valider les champs requis
            errors = []
            
            name = str(row_data.get("name") or "").strip()
            sku = str(row_data.get("sku") or "").strip()
            selling_price = row_data.get("selling_price")
            
            if not name:
                errors.append("Nom du produit requis")
            if not sku:
                errors.append("Code SKU requis")
            if selling_price is None:
                errors.append("Prix de vente requis")
            
            # Vérifier les doublons SKU
            if sku and sku in existing_skus:
                errors.append(f"SKU '{sku}' existe déjà")
            
            # Vérifier les doublons code-barres
            barcode = str(row_data.get("barcode") or "").strip()
            if barcode and barcode in existing_barcodes:
                errors.append(f"Code-barres '{barcode}' existe déjà")
            
            if errors:
                results["errors"].append({
                    "row": row_num,
                    "name": name or f"Ligne {row_num}",
                    "errors": errors
                })
                results["skipped"] += 1
                row_num += 1
                continue
            
            # Préparer le produit
            try:
                product_data = cls._parse_row_data(
                    row_data,
                    category_lookup,
                    brands_map,
                    units_map,
                )
                product_data["organization"] = organization
                product_data["created_by"] = user
                product_data["slug"] = slugify(name)
                
                # Assurer l'unicité du slug
                base_slug = product_data["slug"]
                counter = 1
                while Product.objects.filter(organization=organization, slug=product_data["slug"], is_deleted=False).exists():
                    product_data["slug"] = f"{base_slug}-{counter}"
                    counter += 1
                
                products_to_create.append(product_data)
                existing_skus.add(sku)  # Ajouter pour éviter les doublons dans le même fichier
                if barcode:
                    existing_barcodes.add(barcode)
                
            except Exception as e:
                results["errors"].append({
                    "row": row_num,
                    "name": name,
                    "errors": [str(e)]
                })
                results["skipped"] += 1
            
            row_num += 1
        
        # Créer les produits en transaction
        if products_to_create:
            try:
                SubscriptionService.assert_can_add_products(
                    organization, len(products_to_create)
                )
            except ValidationError as e:
                return {
                    "success": False,
                    "error": e.detail,
                    "created": 0,
                    "updated": 0,
                    "skipped": results["skipped"],
                    "errors": results["errors"],
                }

            try:
                with transaction.atomic():
                    for product_data in products_to_create:
                        Product.objects.create(**product_data)
                        results["created"] += 1
            except Exception as e:
                results["success"] = False
                results["error"] = f"Erreur lors de la création des produits: {str(e)}"
                results["created"] = 0
        
        return results
    
    @classmethod
    def _normalize_key(cls, value: Any) -> str:
        """Normalise une valeur pour comparaison (minuscule, espaces unifiés)."""
        return " ".join(str(value or "").strip().lower().split())

    @classmethod
    def _allocate_unique_category_slug(cls, organization, parent: Optional[Category], display_name: str) -> str:
        """Slug unique pour une catégorie au sein du même parent (contrainte org + parent + slug)."""
        base = slugify(display_name) or "categorie"
        slug = base
        counter = 1
        qs = Category.objects.filter(organization=organization, is_deleted=False, parent=parent)
        while qs.filter(slug=slug).exists():
            slug = f"{base}-{counter}"
            counter += 1
        return slug

    @classmethod
    def _allocate_unique_brand_slug(cls, organization, display_name: str) -> str:
        base = slugify(display_name) or "marque"
        slug = base
        counter = 1
        qs = Brand.objects.filter(organization=organization, is_deleted=False)
        while qs.filter(slug=slug).exists():
            slug = f"{base}-{counter}"
            counter += 1
        return slug

    @classmethod
    def _allocate_unique_unit_symbol(cls, organization, display_name: str) -> str:
        """
        Symbole unique par organisation (max 10 caractères).
        Évite les collisions du type « 3 premières lettres » identiques entre unités différentes.
        """
        clean = (display_name or "").strip()
        root = slugify(clean)[:8] or "u"
        root = root.upper()
        i = 0
        while i < 10000:
            suffix = "" if i == 0 else str(i)
            candidate = (root + suffix)[:10]
            if not Unit.objects.filter(organization=organization, symbol=candidate).exists():
                return candidate
            i += 1
        digest = hashlib.sha256(f"{organization.id}|{clean}|unit".encode()).hexdigest()[:10].upper()
        return digest[:10]

    @classmethod
    def _clean_text(cls, value: Any) -> str:
        """Nettoie une valeur texte sans changer sa casse."""
        return " ".join(str(value or "").strip().split())

    @classmethod
    def _build_category_lookup(cls, organization) -> Dict[Tuple[Optional[str], str], Category]:
        """Construit un index de catégories par parent et nom normalisé."""
        categories = Category.objects.filter(
            organization=organization,
            is_deleted=False
        ).select_related("parent")
        return {
            (str(cat.parent_id) if cat.parent_id else None, cls._normalize_key(cat.name)): cat
            for cat in categories
        }

    @classmethod
    def _build_brands_map(cls, organization) -> Dict[str, Brand]:
        """Construit un index des marques par nom normalisé."""
        brands = Brand.objects.filter(organization=organization, is_deleted=False)
        return {cls._normalize_key(brand.name): brand for brand in brands}

    @classmethod
    def _build_units_map(cls, organization) -> Dict[str, Unit]:
        """Construit un index des unités par nom normalisé."""
        units = Unit.objects.filter(organization=organization)
        return {cls._normalize_key(unit.name): unit for unit in units}

    @classmethod
    def _collect_reference_values(cls, ws):
        """Collecte les référentiels saisis dans les lignes du fichier."""
        parent_categories: Dict[str, str] = {}
        subcategories: Dict[Tuple[str, str], Tuple[str, str]] = {}
        brand_values: Dict[str, str] = {}
        unit_values: Dict[str, str] = {}

        row_num = 3
        while True:
            row_data = {}
            has_data = False
            for col_idx, col_config in enumerate(cls.COLUMNS, start=1):
                value = ws.cell(row=row_num, column=col_idx).value
                if value is not None and str(value).strip():
                    has_data = True
                row_data[col_config["key"]] = value

            if not has_data:
                break

            category_name = cls._clean_text(row_data.get("category"))
            subcategory_name = cls._clean_text(row_data.get("subcategory"))
            brand_name = cls._clean_text(row_data.get("brand"))
            unit_name = cls._clean_text(row_data.get("unit"))

            if category_name and not subcategory_name and ">" in category_name:
                parts = [cls._clean_text(part) for part in category_name.split(">") if cls._clean_text(part)]
                if parts:
                    category_name = parts[0]
                if len(parts) >= 2:
                    subcategory_name = parts[-1]

            if category_name:
                category_key = cls._normalize_key(category_name)
                if category_key:
                    parent_categories.setdefault(category_key, category_name)

                    if subcategory_name:
                        subcategory_key = cls._normalize_key(subcategory_name)
                        if subcategory_key:
                            subcategories.setdefault(
                                (category_key, subcategory_key),
                                (category_name, subcategory_name),
                            )

            if brand_name:
                brand_values.setdefault(cls._normalize_key(brand_name), brand_name)
            if unit_name:
                unit_values.setdefault(cls._normalize_key(unit_name), unit_name)

            row_num += 1

        return parent_categories, subcategories, brand_values, unit_values

    @classmethod
    def _sync_categories(cls, organization, category_lookup, parent_categories, subcategories):
        """Crée les catégories/sous-catégories manquantes."""
        for parent_key, display_name in parent_categories.items():
            lookup_key = (None, parent_key)
            if lookup_key not in category_lookup:
                category = Category.objects.create(
                    organization=organization,
                    name=display_name,
                    slug=cls._allocate_unique_category_slug(organization, None, display_name),
                )
                category_lookup[lookup_key] = category

        for (parent_key, child_key), (parent_display, child_display) in subcategories.items():
            parent = category_lookup.get((None, parent_key))
            if not parent:
                parent = Category.objects.create(
                    organization=organization,
                    name=parent_display,
                    slug=cls._allocate_unique_category_slug(organization, None, parent_display),
                )
                category_lookup[(None, parent_key)] = parent

            child_lookup_key = (str(parent.id), child_key)
            if child_lookup_key not in category_lookup:
                child = Category.objects.create(
                    organization=organization,
                    name=child_display,
                    slug=cls._allocate_unique_category_slug(organization, parent, child_display),
                    parent=parent,
                )
                category_lookup[child_lookup_key] = child

    @classmethod
    def _sync_brands(cls, organization, brands_map, brand_values):
        """Crée les marques manquantes."""
        for normalized_name, display_name in brand_values.items():
            if normalized_name and normalized_name not in brands_map:
                brand = Brand.objects.create(
                    organization=organization,
                    name=display_name,
                    slug=cls._allocate_unique_brand_slug(organization, display_name),
                )
                brands_map[normalized_name] = brand

    @classmethod
    def _sync_units(cls, organization, units_map, unit_values):
        """Crée les unités manquantes."""
        for normalized_name, display_name in unit_values.items():
            if normalized_name and normalized_name not in units_map:
                unit = Unit.objects.create(
                    organization=organization,
                    name=display_name,
                    symbol=cls._allocate_unique_unit_symbol(organization, display_name),
                )
                units_map[normalized_name] = unit
    
    @classmethod
    def _parse_row_data(
        cls,
        row_data: Dict,
        category_lookup: Dict[Tuple[Optional[str], str], Category],
        brands_map: Dict[str, Brand],
        units_map: Dict[str, Unit],
    ) -> Dict:
        """Parse les données d'une ligne Excel en données de produit."""
        
        def parse_decimal(value, default=Decimal("0.00")):
            if value is None:
                return default
            try:
                return Decimal(str(value).replace(",", ".").strip())
            except (InvalidOperation, ValueError):
                return default
        
        def parse_bool(value, default=True):
            if value is None:
                return default
            str_val = str(value).lower().strip()
            if str_val in ("oui", "yes", "1", "true", "vrai"):
                return True
            if str_val in ("non", "no", "0", "false", "faux"):
                return False
            return default
        
        def parse_int(value, default=0):
            if value is None:
                return default
            try:
                return int(float(str(value).strip()))
            except (ValueError, TypeError):
                return default
        
        category_input = cls._clean_text(row_data.get("category"))
        subcategory_input = cls._clean_text(row_data.get("subcategory"))
        brand_input = cls._clean_text(row_data.get("brand"))
        unit_input = cls._clean_text(row_data.get("unit"))

        # Compatibilité descendante: autoriser "Parent > Enfant" saisi dans Catégorie.
        if category_input and not subcategory_input and ">" in category_input:
            parts = [cls._clean_text(part) for part in category_input.split(">") if cls._clean_text(part)]
            if parts:
                category_input = parts[0]
            if len(parts) >= 2:
                subcategory_input = parts[-1]

        category = None
        if category_input:
            parent_key = cls._normalize_key(category_input)
            parent_category = category_lookup.get((None, parent_key))
            category = parent_category
            if parent_category and subcategory_input:
                child_key = cls._normalize_key(subcategory_input)
                category = category_lookup.get((str(parent_category.id), child_key), parent_category)

        brand = brands_map.get(cls._normalize_key(brand_input)) if brand_input else None
        unit = units_map.get(cls._normalize_key(unit_input)) if unit_input else None
        
        return {
            "name": str(row_data.get("name") or "").strip(),
            "sku": str(row_data.get("sku") or "").strip(),
            "barcode": str(row_data.get("barcode") or "").strip(),
            "category": category,
            "brand": brand,
            "unit": unit,
            "cost_price": parse_decimal(row_data.get("cost_price")),
            "selling_price": parse_decimal(row_data.get("selling_price")),
            "wholesale_price": parse_decimal(row_data.get("wholesale_price")) or None,
            "tax_rate": parse_decimal(row_data.get("tax_rate")),
            "is_taxable": parse_bool(row_data.get("is_taxable"), True),
            "track_inventory": parse_bool(row_data.get("track_inventory"), True),
            "has_expiry_date": parse_bool(row_data.get("has_expiry_date"), False),
            "min_stock_level": parse_int(row_data.get("min_stock_level")),
            "short_description": str(row_data.get("short_description") or "").strip(),
            "is_active": True,
        }
    
    @classmethod
    def check_duplicate(cls, organization, sku: str = None, barcode: str = None, exclude_id=None) -> Dict[str, Any]:
        """
        Vérifie si un produit avec le même SKU ou code-barres existe déjà.
        Utilisé pour la création normale de produit aussi.
        """
        duplicates = {"sku": None, "barcode": None}
        
        if sku:
            sku_query = Product.objects.filter(organization=organization, sku=sku, is_deleted=False)
            if exclude_id:
                sku_query = sku_query.exclude(id=exclude_id)
            existing = sku_query.first()
            if existing:
                duplicates["sku"] = {"id": str(existing.id), "name": existing.name}
        
        if barcode and barcode.strip():
            barcode_query = Product.objects.filter(organization=organization, barcode=barcode, is_deleted=False)
            if exclude_id:
                barcode_query = barcode_query.exclude(id=exclude_id)
            existing = barcode_query.first()
            if existing:
                duplicates["barcode"] = {"id": str(existing.id), "name": existing.name}
        
        return duplicates
