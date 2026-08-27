"""
Socle d'export : formatage monétaire, identité de l'émetteur, sous-totaux.

Le premier export du dépôt (le catalogue produits) codait ses décimales en dur à
deux chiffres et n'imprimait que le nom de l'organisation. Ces tests verrouillent
les deux règles que le socle doit tenir à la place.
"""
import io
from decimal import Decimal

from django.test import TestCase
from openpyxl import load_workbook

from apps.core.exports import (
    KIND_MONEY,
    KIND_QUANTITY,
    KIND_TEXT,
    ReportColumn,
    ReportSpec,
    currency_decimals,
    format_number,
    format_quantity,
    organization_identity,
    render_report,
    render_report_pdf,
)
from apps.organizations.models import Organization
from apps.settings.models import Currency


class CurrencyFormattingTests(TestCase):
    """Le CDF n'a pas de décimale ; un `.2f` codé en dur est un bug."""

    def setUp(self):
        Currency.objects.update_or_create(
            code='CDF',
            defaults={'name': 'Franc congolais', 'symbol': 'FC', 'decimal_places': 0},
        )
        Currency.objects.update_or_create(
            code='USD',
            defaults={'name': 'Dollar', 'symbol': '$', 'decimal_places': 2},
        )

    def test_cdf_sans_decimale(self):
        self.assertEqual(currency_decimals('CDF'), 0)
        self.assertEqual(format_number(Decimal('1250036.40'), 0), '1 250 036')

    def test_usd_avec_deux_decimales(self):
        self.assertEqual(currency_decimals('USD'), 2)
        # Convention francophone de bout en bout : l'espace sépare les milliers,
        # la virgule sépare les décimales. Le document imprimé écrivait
        # « 1 250 036.40 » quand l'écran qui le déclenche affiche
        # « 1 250 036,40 » : le même montant se lisait de deux façons.
        self.assertEqual(format_number(Decimal('1250036.40'), 2), '1 250 036,40')

    def test_franc_cfa_aussi_sans_decimale(self):
        # XAF, XOF, JPY, RWF et UGX sont déclarés à 0 par la migration des
        # devises : le repli à 2 ne concerne que les codes absents de la table.
        self.assertEqual(currency_decimals('XAF'), 0)

    def test_devise_absente_de_la_table_replie_sur_deux_decimales(self):
        self.assertEqual(currency_decimals('ZZZ'), 2)

    def test_quantite_sans_zeros_inutiles(self):
        # « 12 » se lit mieux que « 12,000 », mais 0,5 doit rester 0,5.
        self.assertEqual(format_quantity(Decimal('12.000')), '12')
        self.assertEqual(format_quantity(Decimal('0.500')), '0,5')
        self.assertEqual(format_quantity(Decimal('0')), '0')

    def test_milliers_et_decimales_ne_se_confondent_pas(self):
        """
        Le passage en virgule décimale ne doit pas déplacer un séparateur.

        C'est le risque de la conversion : remplacer naïvement « , » puis « . »
        transformerait « 1,234.50 » en « 1.234,50 » ou pire en « 1,234,50 ».
        """
        self.assertEqual(format_number(Decimal('1234.5'), 2), '1 234,50')
        self.assertEqual(format_number(Decimal('1234567.89'), 2), '1 234 567,89')
        self.assertEqual(format_quantity(Decimal('1234.500')), '1 234,5')


class OrganizationIdentityTests(TestCase):
    """Un document imprimé doit dire qui l'émet."""

    def test_identite_complete(self):
        org = Organization.objects.create(
            name='NekaShop', slug='nekashop',
            address='12 avenue du Commerce', city='Kinshasa',
            phone='+243900000000', email='contact@nekashop.cd',
            rccm='CD/KIN/RCCM/22-B-1234', id_nat='01-F4300-N40995A',
            tax_id='A1234567X',
        )
        lines = organization_identity(org)

        self.assertIn('12 avenue du Commerce, Kinshasa', lines)
        self.assertTrue(any('+243900000000' in line for line in lines))
        legal = next(line for line in lines if 'RCCM' in line)
        self.assertIn('ID Nat', legal)
        self.assertIn('NIF', legal)

    def test_champs_absents_ne_laissent_pas_de_ligne_vide(self):
        org = Organization.objects.create(name='Minimal', slug='minimal')
        self.assertEqual(organization_identity(org), [])


class ReportRenderingTests(TestCase):
    """Rendu des deux formats depuis une seule description."""

    def setUp(self):
        Currency.objects.update_or_create(
            code='CDF',
            defaults={'name': 'Franc congolais', 'symbol': 'FC', 'decimal_places': 0},
        )
        self.org = Organization.objects.create(
            name='NekaShop', slug='nekashop', city='Kinshasa',
        )
        self.columns = [
            ReportColumn('name', 'Produit', 40, KIND_TEXT),
            ReportColumn('category', 'Catégorie', 30, KIND_TEXT),
            ReportColumn('quantity', 'Quantité', 20, KIND_QUANTITY),
            ReportColumn('value', 'Valeur', 20, KIND_MONEY),
        ]
        self.rows = [
            {'name': 'Eau 50cl', 'category': 'Boissons',
             'quantity': Decimal('10'), 'value': Decimal('4000')},
            {'name': 'Soda', 'category': 'Boissons',
             'quantity': Decimal('5'), 'value': Decimal('2500')},
            {'name': 'Savon', 'category': 'Hygiène',
             'quantity': Decimal('3'), 'value': Decimal('1200')},
        ]

    def _spec(self, **overrides):
        params = dict(
            title='Test', organization=self.org, columns=self.columns,
            rows=list(self.rows), currency='CDF',
            group_by='category', group_totals=('quantity', 'value'),
        )
        params.update(overrides)
        return ReportSpec(**params)

    def _sheet(self, spec):
        return load_workbook(io.BytesIO(render_report(spec, 'xlsx').getvalue())).active

    def test_les_deux_formats_se_rendent(self):
        spec = self._spec()
        self.assertGreater(len(render_report(spec, 'pdf').getvalue()), 1000)
        self.assertGreater(len(render_report(self._spec(), 'xlsx').getvalue()), 1000)

    def test_format_inconnu_leve(self):
        with self.assertRaises(ValueError):
            render_report(self._spec(), 'docx')

    def test_montants_ecrits_en_nombres(self):
        """Un montant en texte casserait toute somme posée dans le tableur."""
        sheet = self._sheet(self._spec())
        values = [
            cell.value
            for row in sheet.iter_rows()
            for cell in row
            if isinstance(cell.value, (int, float)) and cell.value == 4000
        ]
        self.assertTrue(values, "la valeur 4000 doit exister en nombre")

    def test_sous_totaux_egalent_le_total_general(self):
        sheet = self._sheet(self._spec())
        labels = {}
        for row in sheet.iter_rows():
            first = row[0].value
            if isinstance(first, str) and (
                first.startswith('Sous-total') or first == 'TOTAL GÉNÉRAL'
            ):
                labels[first] = [c.value for c in row]

        self.assertIn('Sous-total Boissons', labels)
        self.assertIn('Sous-total Hygiène', labels)
        self.assertIn('TOTAL GÉNÉRAL', labels)

        # Colonne « Valeur » : 4000 + 2500 = 6500, puis 1200, total 7700.
        self.assertEqual(labels['Sous-total Boissons'][3], 6500)
        self.assertEqual(labels['Sous-total Hygiène'][3], 1200)
        self.assertEqual(labels['TOTAL GÉNÉRAL'][3], 7700)

    def test_pas_de_filtre_auto_sur_un_tableau_groupe(self):
        """Masquer des lignes laisserait des sous-totaux incohérents."""
        self.assertIsNone(self._sheet(self._spec()).auto_filter.ref)

    def test_filtre_auto_couvre_tout_un_tableau_a_plat(self):
        sheet = self._sheet(self._spec(group_by=None, group_totals=()))
        self.assertIsNotNone(sheet.auto_filter.ref)
        last_row = int(sheet.auto_filter.ref.split(':')[1][1:])
        # En-tête + 3 lignes de données, toutes couvertes.
        self.assertEqual(last_row, sheet.max_row)

    def test_rapport_vide_ne_leve_pas(self):
        spec = self._spec(rows=[])
        self.assertGreater(len(render_report(spec, 'pdf').getvalue()), 500)
        self.assertGreater(len(render_report(self._spec(rows=[]), 'xlsx').getvalue()), 500)

    def test_titre_onglet_tronque_et_assaini(self):
        """openpyxl lève au-delà de 31 caractères ou sur `[]:*?/\\`."""
        sheet = self._sheet(self._spec(title='Rapport [stock] / très long ' * 3))
        self.assertLessEqual(len(sheet.title), 31)
        self.assertNotIn('[', sheet.title)
        self.assertNotIn('/', sheet.title)


# =============================================================================
# ALIGNEMENT DU DOCUMENT
# =============================================================================

def _text_left_edges(pdf_bytes):
    """
    Abscisses de tracé du texte d'un PDF reportlab, en millimètres.

    Un `Tm` positionne le texte dans le repère courant, pas dans la page : sans
    rejouer la pile `q` / `Q` / `cm`, tous les paragraphes ressortent à x = 0.
    C'est la seule mesure fiable de l'alignement, un rendu en image ajoutant ses
    propres marges.
    """
    import base64
    import re
    import zlib
    from collections import Counter

    contenu = None
    for match in re.finditer(rb'stream\r?\n', pdf_bytes):
        brut = pdf_bytes[match.end():pdf_bytes.find(b'endstream', match.end())].strip()
        for decoder in (
            lambda b: zlib.decompress(base64.a85decode(b, adobe=True)),
            zlib.decompress,
            lambda b: b,
        ):
            try:
                texte = decoder(brut).decode('latin-1')
            except Exception:
                continue
            if ' Tm' in texte:
                contenu = texte
            break
        if contenu:
            break
    if contenu is None:
        raise AssertionError("aucun flux de contenu lisible dans le PDF")

    jeton = re.compile(r"(-?(?:\d+\.?\d*|\.\d+))|([A-Za-z'\"*]+)")
    pile, ctm, nombres = [], (1.0, 0.0), []
    edges = Counter()
    for match in jeton.finditer(contenu):
        if match.group(1) is not None:
            nombres.append(float(match.group(1)))
            continue
        op = match.group(2)
        if op == 'q':
            pile.append(ctm)
        elif op == 'Q':
            ctm = pile.pop() if pile else (1.0, 0.0)
        elif op == 'cm' and len(nombres) >= 6:
            a, _b, _c, _d, e, _f = nombres[-6:]
            ctm = (ctm[0] * a, ctm[1] + ctm[0] * e)
        elif op in ('Tm', 'Td') and len(nombres) >= 2:
            local = nombres[-6] if op == 'Tm' and len(nombres) >= 6 else nombres[-2]
            edges[round((ctm[1] + ctm[0] * local) / (72 / 25.4), 2)] += 1
        nombres = []
    return edges


class DocumentAlignmentTests(TestCase):
    """
    Le document n'a qu'UN bord gauche.

    Le bandeau d'identité, le titre, le cartouche de synthèse et le tableau
    démarraient à trois abscisses différentes, à un ou deux millimètres près :
    assez peu pour passer deux relectures, assez pour que l'œil lise un bord en
    escalier. Trois causes, toutes invisibles à la lecture du code :
    `Frame` ajoute 6 points de marge intérieure que le calcul de largeur
    ignorait, `Table` retient 6 points de retrait par défaut là où le tableau de
    données en déclarait 3, et `leftIndent` sur un `ParagraphStyle` ne déplace
    pas l'origine de tracé.
    """

    def setUp(self):
        self.org = Organization.objects.create(
            name='NekaShop', slug='nekashop-align',
            address='12 avenue du Commerce', city='Kinshasa',
            phone='+243900000000', email='contact@nekashop.cd',
        )
        self.spec = ReportSpec(
            title='Niveau de stock',
            organization=self.org,
            columns=[
                ReportColumn('produit', 'Produit', 40, KIND_TEXT),
                ReportColumn('quantite', 'Quantité', 20, KIND_QUANTITY),
                ReportColumn('valeur', 'Valeur', 20, KIND_MONEY),
            ],
            rows=[{'produit': 'Eau 50cl', 'quantite': 24, 'valeur': 9600}],
            subtitle='Situation des produits',
            filters_applied=(('Entrepôt', 'Tous'),),
            summary=(('Lignes', '1'), ('Valeur', '9 600,00')),
            group_totals=('valeur',),
        )

    def test_un_seul_bord_gauche_dans_tout_le_document(self):
        edges = _text_left_edges(render_report_pdf(self.spec).getvalue())

        # Le corps du document : identité, titre, synthèse, tableau. On écarte
        # les blocs centrés (signatures) et les colonnes de droite du tableau,
        # qui ont par construction leur propre abscisse.
        gauches = [x for x, n in edges.items() if x < 45 and n >= 3]
        self.assertTrue(gauches, f"aucun bloc de texte à gauche : {dict(edges)}")
        self.assertEqual(
            len(set(gauches)), 1,
            f"le document a {len(set(gauches))} bords gauches : {sorted(set(gauches))} mm",
        )

    def test_leftindent_ne_deplace_pas_un_paragraphe(self):
        """
        Garde contre la « simplification » qui ramènerait le défaut.

        `text_block()` enveloppe le titre dans un tableau, ce qui a tout l'air
        d'un détour inutile : on serait tenté de le remplacer par un
        `leftIndent` sur le style. Ce test montre, sur reportlab lui-même, que
        `leftIndent` laisse l'origine de tracé où elle est - et donc que le
        détour est le seul moyen d'aligner un paragraphe sur une cellule.
        """
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib.units import mm
        from reportlab.platypus import Paragraph, SimpleDocTemplate

        feuilles = getSampleStyleSheet()
        tampon = io.BytesIO()
        SimpleDocTemplate(
            tampon, pagesize=landscape(A4), leftMargin=10 * mm, rightMargin=10 * mm,
        ).build([
            Paragraph('SANS', ParagraphStyle(
                'sans', parent=feuilles['Normal'], leftIndent=0)),
            Paragraph('AVEC', ParagraphStyle(
                'avec', parent=feuilles['Normal'], leftIndent=12)),
        ])

        edges = _text_left_edges(tampon.getvalue())
        self.assertEqual(
            len(edges), 1,
            "reportlab a changé : `leftIndent` déplace désormais le tracé "
            f"({sorted(edges)}). `text_block()` peut alors être simplifié.",
        )
