"""
Socle d'export de rapports : un document se DÉCRIT, deux moteurs le RENDENT.

Le dépôt portait jusqu'ici un seul export serveur, celui des produits, qui
dessinait son PDF et son classeur Excel dans deux fonctions séparées. Le même
piège qu'avec les tickets avant leur refonte : deux tracés à garder synchronisés
dérivent toujours. Ici un rapport se déclare une fois (`ReportSpec`), et
`render_report_pdf` / `render_report_xlsx` en tirent chacun leur rendu.

Deux règles que ce socle fait respecter et que l'export produits enfreint :

1. Les décimales viennent de la devise (`Currency.decimal_places`), jamais d'un
   `.2f` codé en dur. Le CDF n'a pas de décimale : « 1 250 036 FC » et non
   « 1 250 036.40 CDF ».
2. Un document imprimé porte l'identité complète de son émetteur (adresse,
   téléphone, RCCM / ID Nat / NIF), comme les tickets de caisse.
"""
from __future__ import annotations

import io
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from typing import Any, Iterable, Sequence

from django.utils import timezone

# Couleur de marque, alignée sur le frontend (`lib/pdf-utils.ts`).
BRAND_HEX = 'F97316'
HEADER_BG_HEX = 'F97316'
GROUP_BG_HEX = 'FEF3C7'
TOTAL_BG_HEX = 'FDE68A'
MUTED_HEX = '6B7280'
GRID_HEX = 'D1D5DB'

# Types de colonnes. `money` et `quantity` se formatent selon la devise et la
# précision des quantités (3 décimales dans tout le domaine stock).
# Retrait horizontal du texte, en points, à l'intérieur de toute cellule ou
# bloc du document. UNE seule valeur pour l'ensemble : sans elle, le bandeau
# d'identité (retrait 0), le cartouche de synthèse (retrait 6 par défaut de
# reportlab) et le tableau (retrait 3) démarraient à trois abscisses
# différentes, et l'œil lisait un bord gauche en escalier.
CELL_PAD_X = 3

# Marge intérieure du cadre de reportlab. `SimpleDocTemplate` construit son
# `Frame` avec les valeurs par défaut, soit 6 points de chaque côté, et le
# contenu commence donc à `leftMargin + 6`, pas à `leftMargin`. Deux
# conséquences que le code ignorait : le pied de page, tracé directement sur le
# canevas à `leftMargin`, sortait décalé du corps ; et la largeur utile était
# surestimée de 12 points, si bien que chaque tableau débordait légèrement de
# son cadre.
FRAME_PAD_X = 6

KIND_TEXT = 'text'
KIND_NUMBER = 'number'
KIND_QUANTITY = 'quantity'
KIND_MONEY = 'money'
KIND_DATE = 'date'
KIND_DATETIME = 'datetime'
# Quantité exprimée en mots : « 3 PLAQUETTES », « 12 cartons + 3 bouteilles ».
# Le contenu est du texte, mais le lecteur y compare des grandeurs : la colonne
# se range donc avec les nombres, à droite, pour que le bloc chiffré du tableau
# n'ait qu'un seul bord.
KIND_MEASURE = 'measure'

NUMERIC_KINDS = {KIND_NUMBER, KIND_QUANTITY, KIND_MONEY}
# Colonnes qui se rangent à droite : les nombres, et les grandeurs écrites.
RIGHT_ALIGNED_KINDS = NUMERIC_KINDS | {KIND_MEASURE}


@dataclass
class ReportColumn:
    """Une colonne de rapport, décrite une fois pour les deux moteurs."""

    key: str
    header: str
    width: int = 24              # millimètres en PDF, converti pour Excel
    kind: str = KIND_TEXT
    align: str = ''              # vide = déduit du `kind`

    @property
    def effective_align(self) -> str:
        if self.align:
            return self.align
        return 'right' if self.kind in RIGHT_ALIGNED_KINDS else 'left'


@dataclass
class ReportSpec:
    """Description complète d'un rapport, indépendante du format de sortie."""

    title: str
    organization: Any
    columns: Sequence[ReportColumn]
    #: Lignes du rapport. Parcourues une seule fois par rendu ; fournir une
    #: liste si le même `ReportSpec` doit produire les deux formats.
    rows: Iterable[dict]
    subtitle: str = ''
    filters_applied: Sequence[tuple] = field(default_factory=tuple)
    summary: Sequence[tuple] = field(default_factory=tuple)
    group_by: str | None = None
    group_label: str = ''
    group_totals: Sequence[str] = field(default_factory=tuple)
    currency: str = 'CDF'
    landscape_mode: bool = True
    signatures: Sequence[str] = field(default_factory=tuple)
    empty_message: str = 'Aucune donnée pour les critères retenus.'


# --------------------------------------------------------------------------
# Formatage
# --------------------------------------------------------------------------

def currency_decimals(code: str) -> int:
    """
    Nombre de décimales d'une devise, d'après la table `currencies`.

    Le CDF y est déclaré à 0, comme le XAF, le XOF ou le JPY : imprimer
    « 12 500.00 FC » à un commerçant de Kinshasa est à la fois faux et illisible.
    Repli sur 2 pour une devise absente de la table.

    Volontairement sans cache mémoire : la fonction n'est appelée qu'une ou deux
    fois par export, et un cache de portée processus servirait une valeur périmée
    après modification d'une devise, en plus de coupler les tests entre eux
    puisqu'il survivrait au rollback de chacun.
    """
    code = (code or 'CDF').upper()

    from apps.settings.models import Currency

    value = (
        Currency.objects.filter(code=code)
        .values_list('decimal_places', flat=True)
        .first()
    )
    if value is None:
        return 0 if code == 'CDF' else 2
    return value


def currency_symbol(code: str) -> str:
    """Symbole d'affichage d'une devise, replié sur son code."""
    code = (code or 'CDF').upper()
    from apps.settings.models import Currency

    return (
        Currency.objects.filter(code=code)
        .values_list('symbol', flat=True)
        .first()
    ) or code


def to_decimal(value: Any) -> Decimal | None:
    """Convertit en Decimal, ou None si la valeur n'est pas un nombre."""
    if value is None or value == '':
        return None
    if isinstance(value, Decimal):
        return value
    try:
        return Decimal(str(value))
    except (InvalidOperation, TypeError, ValueError):
        return None


def _fr(text: str) -> str:
    """
    Passe une écriture anglo-saxonne en convention francophone.

    « 4,282.60 » devient « 4 282,60 ». Les documents mélangeaient jusqu'ici
    l'espace pour les milliers et le point pour les décimales, alors que les
    écrans qui les déclenchent affichent « 4 282,60 » : le même montant se
    lisait de deux façons selon qu'on le regardait ou qu'on l'imprimait.
    """
    return text.replace(',', '\u202f').replace('.', ',').replace('\u202f', ' ')


def format_number(value: Any, decimals: int) -> str:
    """Formate avec l'espace comme séparateur de milliers (usage francophone)."""
    number = to_decimal(value)
    if number is None:
        return ''
    return _fr(f"{number:,.{decimals}f}")


def format_quantity(value: Any) -> str:
    """
    Quantité au plus juste : 3 décimales possibles, zéros inutiles retirés.

    « 12 » se lit mieux que « 12.000 », mais « 0.5 » doit rester « 0.5 ».
    """
    number = to_decimal(value)
    if number is None:
        return ''
    text = f"{number:,.3f}"
    if '.' in text:
        text = text.rstrip('0').rstrip('.')
    return _fr(text) or '0'


def format_cell(value: Any, kind: str, decimals: int) -> str:
    """Rend une valeur en texte selon le type de colonne."""
    if value is None or value == '':
        return ''
    if kind == KIND_MONEY:
        return format_number(value, decimals)
    if kind == KIND_QUANTITY:
        return format_quantity(value)
    if kind == KIND_MEASURE:
        return str(value)
    if kind == KIND_NUMBER:
        number = to_decimal(value)
        if number is None:
            return str(value)
        return format_number(number, 0 if number == number.to_integral_value() else 2)
    if kind == KIND_DATE:
        return value.strftime('%d/%m/%Y') if isinstance(value, (date, datetime)) else str(value)
    if kind == KIND_DATETIME:
        if isinstance(value, datetime):
            return timezone.localtime(value).strftime('%d/%m/%Y %H:%M') if timezone.is_aware(value) else value.strftime('%d/%m/%Y %H:%M')
        return str(value)
    return str(value)


def organization_identity(organization) -> list[str]:
    """
    Lignes d'identité de l'émetteur, dans l'ordre où un commerçant les lit.

    Même composition que l'en-tête des tickets : le document doit pouvoir servir
    de pièce justificative sans qu'on ait à deviner qui l'a produit.
    """
    lines: list[str] = []

    location = ', '.join(
        part for part in [
            (organization.address or '').replace('\n', ' ').strip(),
            (organization.city or '').strip(),
        ] if part
    )
    if location:
        lines.append(location)

    contact = ' | '.join(
        part for part in [
            f"Tél : {organization.phone}" if organization.phone else '',
            organization.email or '',
        ] if part
    )
    if contact:
        lines.append(contact)

    legal = ' | '.join(
        part for part in [
            f"RCCM : {organization.rccm}" if organization.rccm else '',
            f"ID Nat : {organization.id_nat}" if organization.id_nat else '',
            f"NIF : {organization.tax_id}" if organization.tax_id else '',
        ] if part
    )
    if legal:
        lines.append(legal)

    return lines


def _grouped(spec: ReportSpec):
    """
    Parcourt les lignes en annonçant leur groupe d'appartenance.

    Le regroupement suppose des lignes DÉJÀ triées par la clé de groupe : c'est
    au constructeur du rapport de poser l'`order_by` correspondant, ce qui donne
    les sous-totaux en une seule passe.

    À noter : les constructeurs de `apps.inventory.reports` fournissent des
    listes et non des générateurs, parce que le cartouche de synthèse est écrit
    AVANT le tableau et a besoin des totaux. Le queryset sous-jacent reste lu par
    `iterator(chunk_size=500)`, ce qui borne la mémoire côté base, mais les
    lignes rendues, elles, tiennent en mémoire le temps du rendu.

    Rend des tuples `(nom_du_groupe, ligne)`.
    """
    key = spec.group_by
    for row in spec.rows:
        yield (str(row.get(key) or 'Sans catégorie') if key else None), row


def _accumulate(totals: dict, row: dict, columns: Sequence[str]) -> None:
    """Ajoute une ligne aux totaux courants, en ignorant les valeurs vides."""
    for column_key in columns:
        number = to_decimal(row.get(column_key))
        if number is not None:
            totals[column_key] = totals.get(column_key, Decimal('0')) + number


# --------------------------------------------------------------------------
# Rendu Excel (openpyxl)
# --------------------------------------------------------------------------

def render_report_xlsx(spec: ReportSpec) -> io.BytesIO:
    """
    Rend le rapport en classeur Excel.

    Les montants et quantités sont écrits en NOMBRES, pas en chaînes : un
    gestionnaire qui reçoit le fichier doit pouvoir poser une somme dessus. Le
    format d'affichage porte les décimales de la devise.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    decimals = currency_decimals(spec.currency)
    symbol = currency_symbol(spec.currency)
    money_format = (
        f'#,##0 "{symbol}"' if decimals == 0
        else f'#,##0.{"0" * decimals} "{symbol}"'
    )

    workbook = Workbook()
    sheet = workbook.active
    # Excel refuse les titres d'onglet au-delà de 31 caractères et certains
    # signes de ponctuation : on tronque plutôt que de laisser openpyxl lever.
    sheet.title = ''.join(c for c in spec.title if c not in '[]:*?/\\')[:31] or 'Rapport'

    bold = Font(bold=True)
    white_bold = Font(bold=True, color='FFFFFF')
    muted = Font(size=9, color=MUTED_HEX)
    header_fill = PatternFill('solid', fgColor=HEADER_BG_HEX)
    group_fill = PatternFill('solid', fgColor=GROUP_BG_HEX)
    total_fill = PatternFill('solid', fgColor=TOTAL_BG_HEX)
    thin = Side(style='thin', color=GRID_HEX)
    box = Border(left=thin, right=thin, top=thin, bottom=thin)

    column_count = len(spec.columns)
    row_index = 1

    def write_banner(text: str, font: Font) -> None:
        nonlocal row_index
        cell = sheet.cell(row=row_index, column=1, value=text)
        cell.font = font
        if column_count > 1:
            sheet.merge_cells(
                start_row=row_index, start_column=1,
                end_row=row_index, end_column=column_count,
            )
        row_index += 1

    # En-tête : identité de l'émetteur, puis titre du rapport et contexte.
    write_banner(spec.organization.name, Font(bold=True, size=14, color=BRAND_HEX))
    for line in organization_identity(spec.organization):
        write_banner(line, muted)

    row_index += 1
    write_banner(spec.title, Font(bold=True, size=12))
    if spec.subtitle:
        write_banner(spec.subtitle, muted)
    write_banner(
        f"Généré le {timezone.localtime().strftime('%d/%m/%Y à %H:%M')}", muted
    )
    for label, value in spec.filters_applied:
        write_banner(f"{label} : {value}", muted)

    # Cartouche de synthèse.
    if spec.summary:
        row_index += 1
        for label, value in spec.summary:
            label_cell = sheet.cell(row=row_index, column=1, value=label)
            label_cell.font = bold
            sheet.cell(row=row_index, column=2, value=value)
            row_index += 1

    row_index += 1
    header_row = row_index
    for position, column in enumerate(spec.columns, start=1):
        cell = sheet.cell(row=header_row, column=position, value=column.header)
        cell.font = white_bold
        cell.fill = header_fill
        cell.border = box
        # L'en-tête se range comme sa colonne, exactement comme dans le PDF :
        # les deux moteurs rendent la même description, ils ne doivent pas
        # présenter deux tableaux différents.
        cell.alignment = Alignment(
            horizontal=column.effective_align, vertical='center', wrap_text=True
        )
        sheet.column_dimensions[get_column_letter(position)].width = max(
            10, min(48, column.width * 1.1)
        )
    row_index += 1

    def write_data_row(row: dict, fill: PatternFill | None = None,
                       bold_row: bool = False) -> None:
        nonlocal row_index
        for position, column in enumerate(spec.columns, start=1):
            raw = row.get(column.key)
            if column.kind in NUMERIC_KINDS:
                number = to_decimal(raw)
                # Un montant absent reste vide : un 0 laisserait croire à une
                # valeur nulle mesurée, ce qui fausserait toute moyenne.
                value = float(number) if number is not None else None
            elif column.kind in (KIND_DATE, KIND_DATETIME) and isinstance(raw, datetime):
                value = timezone.localtime(raw).replace(tzinfo=None) if timezone.is_aware(raw) else raw
            elif column.kind == KIND_DATE and isinstance(raw, date):
                value = raw
            else:
                value = format_cell(raw, column.kind, decimals) or None

            cell = sheet.cell(row=row_index, column=position, value=value)
            cell.border = box
            cell.alignment = Alignment(horizontal=column.effective_align)
            if column.kind == KIND_MONEY:
                cell.number_format = money_format
            elif column.kind == KIND_QUANTITY:
                cell.number_format = '#,##0.###'
            elif column.kind == KIND_DATE:
                cell.number_format = 'DD/MM/YYYY'
            elif column.kind == KIND_DATETIME:
                cell.number_format = 'DD/MM/YYYY HH:MM'
            if fill:
                cell.fill = fill
            if bold_row:
                cell.font = bold
        row_index += 1

    def write_totals_row(label: str, totals: dict, fill: PatternFill) -> None:
        nonlocal row_index
        for position, column in enumerate(spec.columns, start=1):
            if position == 1:
                value = label
            elif column.key in totals:
                value = float(totals[column.key])
            else:
                value = None
            cell = sheet.cell(row=row_index, column=position, value=value)
            cell.font = bold
            cell.fill = fill
            cell.border = box
            cell.alignment = Alignment(horizontal=column.effective_align)
            if column.key in totals and column.kind == KIND_MONEY:
                cell.number_format = money_format
            elif column.key in totals and column.kind == KIND_QUANTITY:
                cell.number_format = '#,##0.###'
        row_index += 1

    grand_totals: dict = {}
    group_totals: dict = {}
    current_group = None
    line_count = 0
    last_body_row = header_row

    for group_name, row in _grouped(spec):
        if spec.group_by and group_name != current_group:
            if current_group is not None:
                write_totals_row(f"Sous-total {current_group}", group_totals, group_fill)
            current_group = group_name
            group_totals = {}
            write_data_row({spec.columns[0].key: group_name}, group_fill, bold_row=True)
        write_data_row(row)
        _accumulate(grand_totals, row, spec.group_totals)
        _accumulate(group_totals, row, spec.group_totals)
        line_count += 1
        # Suivi de la dernière ligne réellement écrite : les bandeaux de groupe
        # et les sous-totaux décalent la table, un simple compteur de lignes de
        # données arrêterait le filtre bien avant le bas du tableau.
        last_body_row = row_index - 1

    if spec.group_by and current_group is not None:
        write_totals_row(f"Sous-total {current_group}", group_totals, group_fill)

    if line_count == 0:
        write_banner(spec.empty_message, muted)
    elif spec.group_totals:
        write_totals_row('TOTAL GÉNÉRAL', grand_totals, total_fill)

    # Volet figé : le lecteur garde les en-têtes sous les yeux en faisant défiler
    # plusieurs centaines de lignes.
    sheet.freeze_panes = sheet.cell(row=header_row + 1, column=1)

    # Filtre automatique seulement sur un tableau à plat. Sur un tableau groupé,
    # masquer des lignes laisserait des sous-totaux qui ne correspondent plus à
    # ce qui reste affiché : le lecteur croirait à une incohérence de calcul.
    if line_count and not spec.group_by:
        sheet.auto_filter.ref = (
            f"A{header_row}:{get_column_letter(column_count)}{last_body_row}"
        )

    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


# --------------------------------------------------------------------------
# Rendu PDF (reportlab)
# --------------------------------------------------------------------------

def _numbered_canvas_class():
    """
    Canvas qui n'écrit le pied de page qu'une fois le nombre de pages connu.

    reportlab ne connaît pas le total tant que le document n'est pas bâti : on
    mémorise l'état de chaque page, puis on repasse dessus à l'enregistrement
    pour tamponner « Page X / Y ». Sans ce total, un rapport imprimé ne permet
    pas de vérifier qu'il est complet.
    """
    from reportlab.lib.units import mm
    from reportlab.pdfgen import canvas

    class NumberedCanvas(canvas.Canvas):
        footer_text = ''

        def __init__(self, *args, **kwargs):
            super().__init__(*args, **kwargs)
            self._saved_page_states = []

        def showPage(self):
            self._saved_page_states.append(dict(self.__dict__))
            self._startPage()

        def save(self):
            total = len(self._saved_page_states)
            for state in self._saved_page_states:
                self.__dict__.update(state)
                self._draw_footer(total)
                super().showPage()
            super().save()

        def _draw_footer(self, total: int) -> None:
            # Le pied de page s'aligne sur le même retrait que le corps : posé
            # à la marge nue, il dépassait d'un millimètre à gauche du titre et
            # de la première colonne du tableau.
            width = self._pagesize[0]
            # Le contenu du récit commence à `marge + retrait du cadre`, et son
            # texte 3 points plus loin encore. Le pied de page, tracé sur le
            # canevas nu, doit refaire ce chemin pour tomber sur le même bord.
            inset = 10 * mm + FRAME_PAD_X + CELL_PAD_X
            self.setFont('Helvetica', 7)
            self.setFillColorRGB(0.42, 0.45, 0.5)
            if self.footer_text:
                self.drawString(inset, 8 * mm, self.footer_text)
            self.drawRightString(
                width - inset, 8 * mm,
                f"Page {self._pageNumber} / {total}",
            )

    return NumberedCanvas


def render_report_pdf(spec: ReportSpec) -> io.BytesIO:
    """
    Rend le rapport en PDF A4, en-têtes de colonnes répétés à chaque page.

    Les lignes sont consommées en streaming côté appelant, mais reportlab a
    besoin de la table complète pour la découper : c'est le seul endroit où le
    jeu de données est matérialisé.
    """
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        Image, KeepTogether, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle,
    )

    decimals = currency_decimals(spec.currency)
    symbol = currency_symbol(spec.currency)
    page_size = landscape(A4) if spec.landscape_mode else A4

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=page_size,
        leftMargin=10 * mm,
        rightMargin=10 * mm,
        topMargin=12 * mm,
        bottomMargin=14 * mm,
        title=f"{spec.title} - {spec.organization.name}",
        author=spec.organization.name,
    )

    styles = getSampleStyleSheet()
    org_style = ParagraphStyle(
        'Org', parent=styles['Normal'], fontSize=14, leading=17,
        textColor=colors.HexColor(f'#{BRAND_HEX}'), fontName='Helvetica-Bold',
    )
    # `leftIndent` ne déplace PAS l'origine de tracé d'un paragraphe : mesuré
    # sur le flux PDF, un paragraphe à `leftIndent=3` sort exactement au même
    # x qu'un paragraphe à 0. Le seul retrait qui compte réellement est celui
    # d'une cellule de tableau, d'où `_text_block()` plus bas : tout bloc de
    # texte du document passe par un tableau, donc par le même mécanisme.
    meta_style = ParagraphStyle(
        'Meta', parent=styles['Normal'], fontSize=8, leading=10,
        textColor=colors.HexColor(f'#{MUTED_HEX}'),
    )
    title_style = ParagraphStyle(
        'ReportTitle', parent=styles['Normal'], fontSize=12, leading=15,
        fontName='Helvetica-Bold', spaceBefore=4,
    )
    # Un en-tête doit se ranger comme sa colonne. « Coût unit. » collé à gauche
    # au-dessus de montants collés à droite, c'est le défaut d'alignement qui
    # saute le plus aux yeux sur un tableau large : l'œil ne sait plus quel
    # libellé coiffe quelle colonne de chiffres.
    head_style = ParagraphStyle(
        'Head', parent=styles['Normal'], fontSize=7.5, leading=9,
        fontName='Helvetica-Bold', textColor=colors.white,
    )
    head_right = ParagraphStyle('HeadR', parent=head_style, alignment=2)
    head_center = ParagraphStyle('HeadC', parent=head_style, alignment=1)

    def head(column: ReportColumn):
        align = column.effective_align
        if align == 'right':
            return Paragraph(column.header, head_right)
        if align == 'center':
            return Paragraph(column.header, head_center)
        return Paragraph(column.header, head_style)
    cell_style = ParagraphStyle('Cell', parent=styles['Normal'], fontSize=7.5, leading=9)
    cell_right = ParagraphStyle('CellR', parent=cell_style, alignment=2)
    cell_center = ParagraphStyle('CellC', parent=cell_style, alignment=1)
    bold_cell = ParagraphStyle('CellB', parent=cell_style, fontName='Helvetica-Bold')
    bold_right = ParagraphStyle('CellBR', parent=cell_right, fontName='Helvetica-Bold')

    def styled(text: str, column: ReportColumn, strong: bool = False):
        align = column.effective_align
        if strong:
            return Paragraph(text, bold_right if align == 'right' else bold_cell)
        if align == 'right':
            return Paragraph(text, cell_right)
        if align == 'center':
            return Paragraph(text, cell_center)
        return Paragraph(text, cell_style)

    story: list = []

    # Bandeau d'identité. Le logo est facultatif : une image absente ou illisible
    # (stockage S3 momentanément indisponible) ne doit jamais faire échouer un export.
    header_left: list = [Paragraph(spec.organization.name, org_style)]
    for line in organization_identity(spec.organization):
        header_left.append(Paragraph(line, meta_style))

    logo_cell = ''
    if getattr(spec.organization, 'logo', None):
        try:
            spec.organization.logo.open('rb')
            logo_cell = Image(io.BytesIO(spec.organization.logo.read()),
                              width=22 * mm, height=22 * mm, kind='proportional')
        except Exception:
            logo_cell = ''
        finally:
            try:
                spec.organization.logo.close()
            except Exception:
                pass

    usable_width = page_size[0] - 20 * mm - 2 * FRAME_PAD_X
    header_table = Table(
        [[header_left, logo_cell]],
        colWidths=[usable_width - 26 * mm, 26 * mm],
    )
    header_table.hAlign = 'LEFT'
    header_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
        ('LEFTPADDING', (0, 0), (-1, -1), CELL_PAD_X),
        ('RIGHTPADDING', (0, 0), (-1, -1), CELL_PAD_X),
        ('TOPPADDING', (0, 0), (-1, -1), 0),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
    ]))
    story.append(header_table)
    story.append(Spacer(1, 3 * mm))

    def text_block(flowables: list):
        """
        Enveloppe un bloc de texte dans un tableau sans bordure, pleine largeur.

        C'est le seul moyen fiable de lui donner exactement le même retrait
        qu'une cellule : `leftIndent` ne déplace pas l'origine de tracé, donc un
        paragraphe posé nu dans le récit démarrait 3 points à gauche du bandeau
        d'identité, du cartouche de synthèse et du tableau, tous trois enveloppés
        dans des tableaux. Tout passe désormais par le même mécanisme.
        """
        block = Table([[flowables]], colWidths=[usable_width])
        block.hAlign = 'LEFT'
        block.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('LEFTPADDING', (0, 0), (-1, -1), CELL_PAD_X),
            ('RIGHTPADDING', (0, 0), (-1, -1), CELL_PAD_X),
            ('TOPPADDING', (0, 0), (-1, -1), 0),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
        ]))
        return block

    title_lines: list = [Paragraph(spec.title, title_style)]
    if spec.subtitle:
        title_lines.append(Paragraph(spec.subtitle, meta_style))
    title_lines.append(Paragraph(
        f"Généré le {timezone.localtime().strftime('%d/%m/%Y à %H:%M')}"
        + (f" &nbsp;•&nbsp; Montants en {symbol}" if spec.group_totals else ''),
        meta_style,
    ))
    # Sans le rappel des filtres, un tirage « par catégorie » devient
    # inexploitable dès qu'il quitte l'écran qui l'a produit.
    if spec.filters_applied:
        title_lines.append(Paragraph(
            ' &nbsp;•&nbsp; '.join(f"<b>{label} :</b> {value}"
                                   for label, value in spec.filters_applied),
            meta_style,
        ))
    story.append(text_block(title_lines))
    story.append(Spacer(1, 3 * mm))

    if spec.summary:
        summary_cells = [
            [Paragraph(f"<b>{label}</b><br/><font size=9>{value}</font>", cell_style)
             for label, value in spec.summary]
        ]
        # Largeurs proportionnelles au contenu. À largeur égale, « Période /
        # Tout l'historique » se serrait pendant que « Mouvements / 36 »
        # gardait les trois quarts de sa case vide : le cartouche perdait son
        # rythme et se lisait comme une grille mal remplie. Le plancher de 12 %
        # empêche une case courte de se réduire à un filet.
        weights = [
            max(len(label), len(str(value)) + 2) for label, value in spec.summary
        ]
        floor = 0.12 / len(spec.summary) * sum(weights) if spec.summary else 0
        weights = [max(w, floor) for w in weights]
        total_weight = sum(weights) or 1
        summary_table = Table(
            summary_cells,
            colWidths=[usable_width * w / total_weight for w in weights],
        )
        summary_table.hAlign = 'LEFT'
        summary_table.setStyle(TableStyle([
            ('BOX', (0, 0), (-1, -1), 0.5, colors.HexColor(f'#{GRID_HEX}')),
            ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.HexColor(f'#{GRID_HEX}')),
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#FFF7ED')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), CELL_PAD_X),
            ('RIGHTPADDING', (0, 0), (-1, -1), CELL_PAD_X),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        story.append(summary_table)
        story.append(Spacer(1, 4 * mm))

    # Construction du corps de table, avec sous-totaux intercalés.
    data: list = [[head(c) for c in spec.columns]]
    row_styles: list = []
    grand_totals: dict = {}
    group_totals: dict = {}
    current_group = None
    line_count = 0

    def push_totals(label: str, totals: dict, background: str) -> None:
        cells = []
        for position, column in enumerate(spec.columns):
            if position == 0:
                cells.append(styled(label, column, strong=True))
            elif column.key in totals:
                cells.append(styled(
                    format_cell(totals[column.key], column.kind, decimals),
                    column, strong=True,
                ))
            else:
                cells.append('')
        data.append(cells)
        row_styles.append(('BACKGROUND', (0, len(data) - 1), (-1, len(data) - 1),
                           colors.HexColor(background)))

    for group_name, row in _grouped(spec):
        if spec.group_by and group_name != current_group:
            if current_group is not None:
                push_totals(f"Sous-total {current_group}", group_totals, f'#{GROUP_BG_HEX}')
            current_group = group_name
            group_totals = {}
            header_cells = [''] * len(spec.columns)
            header_cells[0] = Paragraph(f"<b>{group_name}</b>", bold_cell)
            data.append(header_cells)
            row_styles.append(('BACKGROUND', (0, len(data) - 1), (-1, len(data) - 1),
                               colors.HexColor(f'#{GROUP_BG_HEX}')))
            row_styles.append(('SPAN', (0, len(data) - 1), (-1, len(data) - 1)))

        data.append([
            styled(format_cell(row.get(c.key), c.kind, decimals) or '-', c)
            for c in spec.columns
        ])
        _accumulate(grand_totals, row, spec.group_totals)
        _accumulate(group_totals, row, spec.group_totals)
        line_count += 1

    if spec.group_by and current_group is not None:
        push_totals(f"Sous-total {current_group}", group_totals, f'#{GROUP_BG_HEX}')
    if line_count and spec.group_totals:
        push_totals('TOTAL GÉNÉRAL', grand_totals, f'#{TOTAL_BG_HEX}')

    if line_count == 0:
        story.append(text_block([Paragraph(spec.empty_message, meta_style)]))
    else:
        total_declared = sum(c.width for c in spec.columns) or 1
        col_widths = [usable_width * c.width / total_declared for c in spec.columns]
        table = Table(data, colWidths=col_widths, repeatRows=1)
        # `hAlign` vaut CENTER par défaut : si la somme des largeurs tombait un
        # demi-millimètre sous la largeur utile, tout le tableau se décalerait
        # d'un quart de millimètre sans que rien ne le signale.
        table.hAlign = 'LEFT'
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(f'#{HEADER_BG_HEX}')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor(f'#{GRID_HEX}')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1),
             [colors.white, colors.HexColor('#FAFAFA')]),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
            ('LEFTPADDING', (0, 0), (-1, -1), CELL_PAD_X),
            ('RIGHTPADDING', (0, 0), (-1, -1), CELL_PAD_X),
        ] + row_styles))
        story.append(table)

    if spec.signatures:
        story.append(Spacer(1, 10 * mm))
        signature_cells = [[
            Paragraph(f"{label}<br/><br/><br/>______________________", cell_center)
            for label in spec.signatures
        ]]
        signature_table = Table(
            signature_cells,
            colWidths=[usable_width / len(spec.signatures)] * len(spec.signatures),
        )
        signature_table.hAlign = 'LEFT'
        signature_table.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('LEFTPADDING', (0, 0), (-1, -1), CELL_PAD_X),
            ('RIGHTPADDING', (0, 0), (-1, -1), CELL_PAD_X),
        ]))
        story.append(KeepTogether(signature_table))

    canvas_class = _numbered_canvas_class()
    canvas_class.footer_text = f"{spec.organization.name} - {spec.title}"
    doc.build(story, canvasmaker=canvas_class)
    buffer.seek(0)
    return buffer


RENDERERS = {
    'pdf': render_report_pdf,
    'xlsx': render_report_xlsx,
}

CONTENT_TYPES = {
    'pdf': 'application/pdf',
    'xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
}


def render_report(spec: ReportSpec, fmt: str) -> io.BytesIO:
    """Rend le rapport dans le format demandé (`pdf` ou `xlsx`)."""
    try:
        return RENDERERS[fmt](spec)
    except KeyError:
        raise ValueError(f"Format d'export inconnu : {fmt}")
